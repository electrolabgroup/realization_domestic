import tkinter as tk
from tkinter import scrolledtext
import threading
from datetime import datetime, timedelta
import time
import os
from openpyxl import load_workbook, Workbook


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("C&R Updater")
        self.root.geometry("900x280")

        # Load the icon image
        icon_path = "ELEC.png"
        icon_img = tk.PhotoImage(file=icon_path)

        # Set window icon
        self.root.tk.call('wm', 'iconphoto', self.root._w, icon_img)

        # Create text area to display output
        self.output_text = scrolledtext.ScrolledText(root, height=10, width=90, font=("Arial", 12))
        self.output_text.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

        # Create Start button
        self.start_btn = tk.Button(root, text="Start", command=self.start_processing, font=("Arial", 12), bg="#4CAF50",
                                   fg="white")
        self.start_btn.grid(row=1, column=0, padx=10, pady=10, sticky="ew")

        # Create Stop button
        self.stop_btn = tk.Button(root, text="Stop", command=self.stop_processing, font=("Arial", 12), bg="#f44336",
                                  fg="white", state=tk.DISABLED)
        self.stop_btn.grid(row=1, column=1, padx=10, pady=10, sticky="ew")

        # Create Exit button
        self.exit_btn = tk.Button(root, text="Exit", command=root.quit, font=("Arial", 12))
        self.exit_btn.grid(row=1, column=2, padx=10, pady=10, sticky="ew")

        # Variable to control processing loop
        self.processing = False

    def start_processing(self):
        self.processing = True
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        # Start processing in a separate thread
        threading.Thread(target=self.process_data).start()

    def stop_processing(self):
        self.processing = False
        self.start_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)
        self.log_output("Processing stopped.")

    def process_data(self):

        while self.processing:

            # !/usr/bin/env python
            # coding: utf-8

            # In[1]:

            # !/usr/bin/env python
            # coding: utf-8

            # In[1]:

            start_time = time.time()
            self.clear_text_area()

            self.log_output(f"Calculating realization!")
            from datetime import datetime

            import requests
            import pandas as pd

            # Define the base URL and endpoint for sales orders
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Sales Order'
            url = base_url + endpoint

            # Define the headers for the request
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Create a session with retries
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry

            session = requests.Session()
            retry = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount('http://', adapter)
            session.mount('https://', adapter)

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Define the parameters for the request
                params = {
                    'fields': '["name","customer","freight_amt","payment_terms_template","commission_rate","packing_charges","oem_discount","items.item_code","items.qty","items.amount","items.item_name","items.jars_specification"]',
                    'limit_start': limit_start,
                    'limit_page_length': limit_page_length,
                    'filters': '[["reason_for_hold", "like", "Realization%"], ["status", "not in", ["Cancelled","Draft"]],["name", "not like", "%SOEXP%"],["name", "not like", "%ROYALTY%"]]'
                }

                try:
                    response = session.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        so_data = data['data']
                        if not so_data:
                            break  # No more data to fetch
                        all_data.extend(so_data)
                        limit_start += limit_page_length
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Check if any data was fetched and create DataFrame
            if all_data:
                so_df = pd.json_normalize(all_data)
            else:
                columns = ["name", "customer", "freight_amt", "payment_terms_template", "commission_rate",
                           "packing_charges", "oem_discount", "item_code", "qty", "amount", "item_name",
                           "jars_specification"]
                so_df = pd.DataFrame(columns=columns)

            import requests
            import pandas as pd

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Item Price'
            url = base_url + endpoint

            # Define the headers for the request
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Create a session with retries
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry

            session = requests.Session()
            retry = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount('http://', adapter)
            session.mount('https://', adapter)

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Define the parameters for the request
                params = {
                    'fields': '["item_code","price_list_rate"]',
                    'limit_start': limit_start,
                    'limit_page_length': limit_page_length,
                    'filters': '[["price_list", "=", "Standard Spares"]]'
                }

                try:
                    response = session.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        spares_data = data['data']
                        if not spares_data:
                            break  # No more data to fetch
                        all_data.extend(spares_data)
                        limit_start += limit_page_length
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Convert the collected data to a DataFrame
            spares_df = pd.json_normalize(all_data)

            spares_df['price_list_rate'] = spares_df['price_list_rate'] * 0.80

            import requests
            import pandas as pd

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Item Price'
            url = base_url + endpoint

            # Define the headers for the request
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Create a session with retries
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry

            session = requests.Session()
            retry = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount('http://', adapter)
            session.mount('https://', adapter)

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Define the parameters for the request
                params = {
                    'fields': '["item_code","price_list_rate"]',
                    'limit_start': limit_start,
                    'limit_page_length': limit_page_length,
                    'filters': '[["price_list", "=", "Realization Target Machine 24_25"]]'
                }

                try:
                    response = session.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        machine_data_r = data['data']
                        if not machine_data_r:
                            break  # No more data to fetch
                        all_data.extend(machine_data_r)
                        limit_start += limit_page_length
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Convert the collected data to a DataFrame
            machine_df_r = pd.json_normalize(all_data)

            import requests
            import pandas as pd

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Item Price'
            url = base_url + endpoint

            # Define the headers for the request
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Create a session with retries
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry

            session = requests.Session()
            retry = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount('http://', adapter)
            session.mount('https://', adapter)

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Define the parameters for the request
                params = {
                    'fields': '["item_code","price_list_rate"]',
                    'limit_start': limit_start,
                    'limit_page_length': limit_page_length,
                    'filters': '[["price_list", "=", "MRP List price Machine 24_25"]]'
                }

                try:
                    response = session.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        machine_data_mrp = data['data']
                        if not machine_data_mrp:
                            break  # No more data to fetch
                        all_data.extend(machine_data_mrp)
                        limit_start += limit_page_length
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Convert the collected data to a DataFrame
            machine_mrp = pd.json_normalize(all_data)

            # Display the first few rows of the DataFrame
            machine_mrp.shape
            machine_mrp.rename(columns={'price_list_rate': 'mrp_price'}, inplace=True)

            # In[6]:

            machine_df = pd.merge(machine_mrp, machine_df_r, on='item_code', how='inner')
            machine_df = machine_df.round(2)
            machine_df.head()

            # In[ ]:

            so_df.fillna(0, inplace=True)

            # In[7]:

            # Function to replace values based on substring match
            def replace_values(row):
                # Condition 2: Drop rows where 'Item Name' contains "IQ/OQ" or "IQOQ" and 'Amount' is 0
                if ('IQ/OQ' in row['item_name'] or 'IQOQ' in row['item_name']) and row['amount'] == 0:
                    so_df.drop(index=row.name, inplace=True)

                if 'Kloudface' in row['item_name'] and 'Model' in row['item_name']:
                    so_df.loc[row.name, 'item_name'] = 'Kloudface Backup All EKF Model'
                    so_df.loc[
                        row.name, 'item_name'] = '000001520800'  # Assuming 'Item Code' is a string, adjust accordingly

            # Apply the function to each row"
            so_df.apply(replace_values, axis=1)

            # In[9]:

            # In[11]:

            # Concatenate the two DataFrames along the rows
            df = pd.concat([machine_df, spares_df], ignore_index=True)

            df.drop_duplicates(subset='item_code', keep='first', inplace=True)
            # df = spares_df.copy()

            # In[12]:

            rl_df = pd.merge(so_df, df, on=['item_code'], how='left')

            # In[13]:

            # In[14]:

            # First, convert any empty strings or NaN values in 'price_list_rate' column to None
            rl_df['price_list_rate'] = rl_df['price_list_rate'].replace('', pd.NA).astype('float64')

            # Group the DataFrame by 'name' and filter out groups where any 'price_list_rate' is missing
            rl_df = rl_df.groupby('name').filter(lambda x: not x['price_list_rate'].isnull().any())

            # In[18]:

            # Create a DataFrame with non-matched Item Codes in so_df
            merged_df = pd.merge(so_df, df, on=['item_code'], how='left', indicator=True)
            sheet_df3 = merged_df[merged_df['_merge'] == 'left_only'].drop(columns=['_merge'])

            # In[21]:

            sheet_df3.rename(columns={'name': 'Sales Order'}, inplace=True)
            sheet_df3 = sheet_df3[
                ['Sales Order', 'customer', 'item_code', 'item_name', 'freight_amt', 'packing_charges', 'oem_discount',
                 'qty', 'amount']]

            # In[22]:

            sheet_df3['Date and Time'] = datetime.now()
            # In[25]:

            item_dissolution_df = pd.read_excel('code_file.xlsx')

            # In[26]:

            item_dissolution_df.head()

            # In[27]:

            rl_df = rl_df[rl_df['amount'] >= 0]
            # rl_df = rl_df[~((rl_df['Item Name_y'].str.startswith('Item Name_y')) & (rl_df['Amount'] == 0))]

            # In[28]:

            # Define the mapping dictionary
            p_tdf = pd.read_excel('Payment_Terms_Template.xlsx')
            # Convert the DataFrame to a dictionary
            payment_term_mapping = dict(zip(p_tdf['Name'], p_tdf['Order Process Percentage'] * 100))

            # In[29]:

            rl_df = rl_df[rl_df['payment_terms_template'] != 'Select']
            # Replace values in the 'Payment Term' column using the mapping dictionary
            rl_df['payment_terms_template'] = rl_df['payment_terms_template'].replace(payment_term_mapping)

            def replace_target_pandas(rl_df, item_dissolution_df):
                # Create a mask to identify rows in rl_df that match the dissolution item codes
                mask = rl_df['item_code'].isin(item_dissolution_df['Dissolution Item Code'])

                # Apply adjustments only to the filtered rows
                for idx, row in rl_df[mask].iterrows():
                    jar_specification = row['jars_specification']
                    item_name = row['item_name']
                    adjustment = 0  # Initialize adjustment to 0

                    if 'Inspire-8' in item_name:
                        adjustment = {
                            'Glass-Clear': 11400,
                            'All Glass-Clear': 15200,
                            'Glass Amber': 17620,
                            'All Glass-Amber': 21760,
                            'Merlon Amber': 5200,
                        }.get(jar_specification, 0)

                    elif 'Inspire-14' in item_name:
                        adjustment = {
                            'Glass-Clear': 22800,
                            'All Glass-Clear': 26600,
                            'Glass Amber': 33940,
                            'All Glass-Amber': 38080,
                            'Merlon Amber': 9100,
                        }.get(jar_specification, 0)

                    elif 'TrustE-8' in item_name:
                        adjustment = {
                            'All Glass-Clear': 3800,
                            'Glass Amber': 11500,
                            'All Glass-Amber': 17400,
                            'Merlon Amber': -6200,
                            'Merlon Clear': -11400,
                        }.get(jar_specification, 0)

                    elif 'Tablet Dissolution Tester Model TrustE-14' in item_name:
                        adjustment = {
                            'All Glass-Clear': 3800,
                            'Glass Amber': 21700,
                            'All Glass-Amber': 27600,
                            'Merlon Amber': -13700,
                            'Merlon Clear': -22800,
                        }.get(jar_specification, 0)

                    # Adjust MRP based on the conditions for the specific row
                    rl_df.at[idx, 'mrp_price'] += adjustment
                    rl_df['price_list_rate'] == rl_df['mrp_price'] * 0.80

                return rl_df

            # Call the function to replace values in rl_df
            rl_df = replace_target_pandas(rl_df, item_dissolution_df)

            def replace_target_realization_target(rl_df, item_dissolution_df):
                # Create a mask to identify rows in rl_df that match the dissolution item codes
                mask = rl_df['item_code'].isin(item_dissolution_df['Dissolution Item Code'])

                # Apply adjustments only to the filtered rows
                for idx, row in rl_df[mask].iterrows():
                    # Update the 'Realization Target' column with 80% of 'MRP'
                    rl_df.at[idx, 'Realization Target'] = row['mrp_price'] * 0.80

                return rl_df

            rl_df = replace_target_realization_target(rl_df, item_dissolution_df)

            # In[42]:

            if (rl_df['oem_discount'] == 25).any():
                rl_df['price_list_rate'] = rl_df['mrp_price'] * 0.75
                # Convert 'Freight Amt' to numeric, replacing non-numeric values with NaN
                rl_df['freight_amt'] = pd.to_numeric(rl_df['freight_amt'], errors='coerce')
                rl_df['target'] = rl_df['price_list_rate'] * rl_df['qty']

            elif (rl_df['oem_discount'] == 0).any():
                # Convert 'Freight Amt' to numeric, replacing non-numeric values with NaN
                rl_df['freight_amt'] = pd.to_numeric(rl_df['freight_amt'], errors='coerce')
                # rl_df = rl_df[['Name','Freight Amt','Amount','Total Commission','Payment Term','Quantity','Realization Target']]
                rl_df['target'] = rl_df['price_list_rate'] * rl_df['qty']

            else:
                rl_df['target'] = ' '

            # In[43]:

            # Convert 'Realization Target' column to integer
            rl_df['target'] = pd.to_numeric(rl_df['target'], errors='coerce').fillna(0).astype(
                int)
            rl_df['freight_amt'] = pd.to_numeric(rl_df['freight_amt'], errors='coerce').fillna(0).astype(int)
            rl_df['amount'] = pd.to_numeric(rl_df['amount'], errors='coerce').fillna(0).astype(int)
            rl_df['commission_rate'] = pd.to_numeric(rl_df['commission_rate'], errors='coerce').fillna(0).astype(int)
            rl_df['payment_terms_template'] = pd.to_numeric(rl_df['payment_terms_template'], errors='coerce').fillna(
                0).astype(int)
            rl_df['qty'] = pd.to_numeric(rl_df['qty'], errors='coerce').fillna(0).astype(int)

            # Identify Names with at least one Amount zero
            names_with_zero_amount = rl_df.loc[rl_df['amount'] == 0, 'name'].unique()

            # In[46]:

            # Identify Names with multiple non-zero values
            names_with_multiple_non_zero = rl_df.loc[
                (rl_df['amount'] != 0) & (rl_df.duplicated(subset='name', keep=False)), 'name'].unique()

            # In[47]:

            # Create DataFrame for names with at least one Amount zero
            df_with_zero = rl_df[rl_df['name'].isin(names_with_zero_amount)].copy()

            # In[48]:

            # Create DataFrame for names with only non-zero values
            df_non_zero = rl_df[~rl_df['name'].isin(names_with_zero_amount)].copy()

            # In[53]:

            # Group by 'Name' and apply aggregation functions for names with at least one Amount zero
            aggregation_functions_with_zero = {
                'freight_amt': 'mean',
                'amount': 'sum',
                'commission_rate': 'first',
                'payment_terms_template': 'first',
                'target': 'sum',
                'customer': 'first'
            }

            df_with_zero = df_with_zero.groupby('name').agg(aggregation_functions_with_zero).reset_index()

            # In[56]:

            # Concatenate the DataFrames
            rl_df = pd.concat([df_non_zero, df_with_zero], ignore_index=True)

            # In[59]:

            # Calculate the count of line items per Sales Order
            line_item_count = rl_df.groupby('name').size().reset_index(name='Line Item Count')
            rl_df = pd.merge(rl_df, line_item_count, on='name', how='left')

            # In[64]:

            # Calculate the average freight amount per Sales Order by dividing total freight by line item count
            rl_df['Average Freight'] = rl_df['freight_amt'] / rl_df['Line Item Count']

            # Drop unnecessary columns if needed
            rl_df = rl_df.drop(['Line Item Count'], axis=1)

            # In[ ]:

            import requests
            import pandas as pd

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Customer'
            url = base_url + endpoint

            # Define the headers for the request
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Create a session with retries
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry

            session = requests.Session()
            retry = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
            )
            adapter = HTTPAdapter(max_retries=retry)
            session.mount('http://', adapter)
            session.mount('https://', adapter)

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Define the parameters for the request
                params = {
                    'fields': '["name","customer_name","custom_credit_rate"]',
                    'limit_start': limit_start,
                    'limit_page_length': limit_page_length
                }

                try:
                    response = session.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        customer_data = data['data']
                        if not customer_data:
                            break  # No more data to fetch
                        all_data.extend(customer_data)
                        limit_start += limit_page_length
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Convert the collected data to a DataFrame
            cr_df = pd.json_normalize(all_data)

            # In[ ]:

            # Replace 'NONE' with 0
            cr_df['custom_credit_rate'] = cr_df['custom_credit_rate'].replace('NONE', 0)

            # Convert to numeric and replace non-numeric values with 0
            cr_df['custom_credit_rate'] = pd.to_numeric(cr_df['custom_credit_rate'], errors='coerce').fillna(0)

            # In[ ]:

            cr_df.rename(columns={'name': 'customer'}, inplace=True)

            # In[ ]:

            fl_df = pd.merge(rl_df, cr_df, on='customer', how='outer')
            # Set NaN values in 'Credit Rate' column to zero
            fl_df['custom_credit_rate'] = fl_df['custom_credit_rate'].fillna(0)

            # In[ ]:

            # Assuming Credit Rate is a column in fl_df
            fl_df['payment_terms_template'] = fl_df.apply(
                lambda row: row['payment_terms_template'] if row['custom_credit_rate'] == 0 else 100 - row[
                    'custom_credit_rate'], axis=1)

            # In[ ]:

            # Ensure no NaN values in 'amount', 'commission_rate', 'Average Freight', and 'payment_terms_template'
            fl_df['amount'] = fl_df['amount'].fillna(0)
            fl_df['commission_rate'] = fl_df['commission_rate'].fillna(0)
            fl_df['Average Freight'] = fl_df['Average Freight'].fillna(0)
            fl_df['payment_terms_template'] = fl_df['payment_terms_template'].fillna(
                1)  # Default to 1 if NaN, assuming it's a multiplier


            # Calculate the Realization Value
            fl_df['realization_value'] = (
                    (fl_df['amount'] * (1 - fl_df['commission_rate'] / 100) - fl_df['Average Freight']) * fl_df[
                'payment_terms_template']
            )

            # rl_df = rl_df['Payment Term']

            fl_df['realization_value'] = fl_df['realization_value'].fillna(0)
            fl_df['target'] = fl_df['target'].fillna(0)

            fl_df['Realization line item'] = (fl_df['realization_value'] / fl_df['target']).round(1)

            # In[ ]:

            fl_df['realization_value'] = fl_df['realization_value'].fillna(0)
            fl_df['target'] = fl_df['target'].fillna(0)

            fl_df['realization_value_sum'] = (fl_df.groupby('name')['realization_value'].transform('sum'))
            fl_df['target_sum'] = (fl_df.groupby('name')['target'].transform('sum'))

            fl_df['realization'] = (fl_df['realization_value_sum'] / fl_df['target_sum']).round(1)

            # In[ ]:

            # Drop rows where the 'Name' column is null
            fl_df = fl_df.dropna(subset=['name'])

            # In[ ]:

            # Assuming 'Realization in %' is a numeric column
            fl_df['realization'] = pd.to_numeric(fl_df['realization'], errors='coerce')

            ## Assuming fl_df is your DataFrame
            if not fl_df.empty:
                fl_df['realization'] = fl_df['realization'].apply(lambda x: max(x, 0))

                # Check if DataFrame is still not empty after applying operations
                if not fl_df.empty:
                    # Updating 'Order Status' based on 'Realization in %'
                    fl_df.loc[fl_df['realization'] == 0, 'reason_for_hold'] = 'Pending for Realization%'
                    fl_df.loc[(fl_df['realization'] > 0) & (
                            fl_df['realization'] < 100), 'reason_for_hold'] = 'Pending for approval'
                    fl_df.loc[fl_df['realization'] >= 100, 'reason_for_hold'] = 'Order Review Pending'
            else:
                fl_df['reason_for_hold'] = ' '
                self.log_output("DataFrame is empty. Ignoring update operation.")


            sheet_df1 = fl_df[
                ['name', 'item_code', 'item_name', 'freight_amt', 'customer', 'qty', 'amount', 'commission_rate',
                 'packing_charges', 'payment_terms_template', 'realization_value', 'target', 'Realization line item',
                 'realization', 'reason_for_hold']]



            sheet_df1.rename(columns={"name": "Sales Order"}, inplace=True)



            sheet_df2 = fl_df[['name', 'customer', 'realization_value', 'target', 'realization', 'reason_for_hold']]

            # In[ ]:

            sheet_df1['Date and Time'] = datetime.now()

            # In[ ]:

            sheet_df1['realization'] = sheet_df1['realization'] / 100

            # In[ ]:

            aggregation = {
                'realization_value': 'sum',
                'target': 'sum',
                'realization': 'first',
                'reason_for_hold': 'first'
            }

            # In[ ]:

            sheet_df2 = sheet_df2.groupby('name').agg(aggregation).reset_index()

            # In[ ]:

            sheet_df2['realization_value'] = sheet_df2['realization_value'] / 100

            # In[ ]:

            sheet_df2['Date and Time'] = datetime.now()

            # In[ ]:

            sheet_df3.head()

            # In[ ]:

            selected_columns = ['name', 'realization', 'reason_for_hold']
            fl_df1 = fl_df[selected_columns]
            fl_df1.rename(columns={"name": "name"}, inplace=True)
            # Drop duplicates based on the 'name' column
            fl_df1 = fl_df1.drop_duplicates(subset=['name'])
            fl_df1.head()

            # In[ ]:

            selected_columns = ['item_code', 'Realization line item']
            fl_df2 = fl_df[selected_columns]
            fl_df2.rename(columns={
                "Realization line item": 'realization'}, inplace=True)

            # In[ ]:

            fl_df3 = fl_df[['name', 'realization', 'Realization line item', 'reason_for_hold']]

            # In[ ]:

            condition1 = sheet_df1['Sales Order'].str.contains('SODM')
            condition2 = sheet_df1['Sales Order'].str.contains('SOEXP')
            condition3 = sheet_df1['Sales Order'].str.contains('SODS')

            machine_sheet_df1 = sheet_df1[~(condition2 | condition3)]  # Remove rows with SOEXP or SODS
            spares_sheet_df1 = sheet_df1[~(condition1 | condition2)]  # Remove rows with SODM or SOEXP
            export_sheet_df1 = sheet_df1[~(condition1 | condition3)]  # Remove rows with SODM or SODS

            condition1 = sheet_df2['name'].str.contains('SODM')
            condition2 = sheet_df2['name'].str.contains('SOEXP')
            condition3 = sheet_df2['name'].str.contains('SODS')

            machine_sheet_df2 = sheet_df2[~(condition2 | condition3)]  # Remove rows with SOEXP or SODS
            spares_sheet_df2 = sheet_df2[~(condition1 | condition2)]  # Remove rows with SODM or SOEXP
            export_sheet_df2 = sheet_df2[~(condition1 | condition3)]  # Remove rows with SODM or SODS

            condition1 = sheet_df3['Sales Order'].str.contains('SODM')
            condition2 = sheet_df3['Sales Order'].str.contains('SOEXP')
            condition3 = sheet_df3['Sales Order'].str.contains('SODS')

            machine_sheet_df3 = sheet_df3[~(condition2 | condition3)]  # Remove rows with SOEXP or SODS
            spares_sheet_df3 = sheet_df3[~(condition1 | condition2)]  # Remove rows with SODM or SOEXP
            export_sheet_df3 = sheet_df3[~(condition1 | condition3)]  # Remove rows with SODM or SODS

            # In[ ]:

            # Get the current month and year
            current_month_year = datetime.now().strftime("%Y-%m")

            # Define the file path for the new Excel file for export data
            machinepath = f'Data\MachineData_{current_month_year}.xlsx'

            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(machinepath), exist_ok=True)

            # Function to append a DataFrame to a specific sheet in the Excel file for export data
            def append_to_excel(df, sheet_name):
                # Load the existing Excel file
                try:
                    book = load_workbook(machinepath)
                except FileNotFoundError:
                    # If the file doesn't exist, create a new one
                    book = Workbook()

                # Get the active worksheet
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    startrow = sheet.max_row + 1  # Start from the next row
                else:
                    sheet = book.create_sheet(title=sheet_name)
                    startrow = 1  # Start from the first row

                # Append the DataFrame to the sheet
                data = df.values.tolist()

                # If the sheet is empty, write column headers first
                if startrow == 1:
                    headers = list(df.columns)
                    sheet.append(headers)
                    startrow += 1

                for row in data:
                    sheet.append(row)

                # Save the changes
                book.save(machinepath)

            # DataFrames and their corresponding sheet names for export data
            machine_dataframes_and_sheets = {
                'Relization Item Lines': machine_sheet_df1,
                'Relization Mean': machine_sheet_df2,
                'Missing Items': machine_sheet_df3
            }

            # Loop through the dictionary and append each DataFrame to its corresponding sheet for export data
            for sheet_name, df in machine_dataframes_and_sheets.items():
                append_to_excel(df, sheet_name)

            # In[ ]:

            # Get the current month and year
            current_month_year = datetime.now().strftime("%Y-%m")

            # Define the file path for the new Excel file for export data
            sparepath = f'Data\SparesData_{current_month_year}.xlsx'

            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(sparepath), exist_ok=True)

            # Function to append a DataFrame to a specific sheet in the Excel file for export data
            def append_to_excel(df, sheet_name):
                # Load the existing Excel file
                try:
                    book = load_workbook(sparepath)
                except FileNotFoundError:
                    # If the file doesn't exist, create a new one
                    book = Workbook()

                # Get the active worksheet
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    startrow = sheet.max_row + 1  # Start from the next row
                else:
                    sheet = book.create_sheet(title=sheet_name)
                    startrow = 1  # Start from the first row

                # Append the DataFrame to the sheet
                data = df.values.tolist()

                # If the sheet is empty, write column headers first
                if startrow == 1:
                    headers = list(df.columns)
                    sheet.append(headers)
                    startrow += 1

                for row in data:
                    sheet.append(row)

                # Save the changes
                book.save(sparepath)

            # DataFrames and their corresponding sheet names for export data
            spare_dataframes_and_sheets = {
                'Relization Item Lines': spares_sheet_df1,
                'Relization Mean': spares_sheet_df2,
                'Missing Items': spares_sheet_df3
            }

            # Loop through the dictionary and append each DataFrame to its corresponding sheet for export data
            for sheet_name, df in spare_dataframes_and_sheets.items():
                append_to_excel(df, sheet_name)

            # In[ ]:

            # Get the current month and year
            current_month_year = datetime.now().strftime("%Y-%m")

            # Define the file path for the new Excel file for export data
            exportpath = f'Data\ExportData_{current_month_year}.xlsx'

            # Create the directory if it doesn't exist
            os.makedirs(os.path.dirname(exportpath), exist_ok=True)

            # Function to append a DataFrame to a specific sheet in the Excel file for export data
            def append_to_excel(df, sheet_name):
                # Load the existing Excel file
                try:
                    book = load_workbook(exportpath)
                except FileNotFoundError:
                    # If the file doesn't exist, create a new one
                    book = Workbook()

                # Get the active worksheet
                if sheet_name in book.sheetnames:
                    sheet = book[sheet_name]
                    startrow = sheet.max_row + 1  # Start from the next row
                else:
                    sheet = book.create_sheet(title=sheet_name)
                    startrow = 1  # Start from the first row

                # Append the DataFrame to the sheet
                data = df.values.tolist()

                # If the sheet is empty, write column headers first
                if startrow == 1:
                    headers = list(df.columns)
                    sheet.append(headers)
                    startrow += 1

                for row in data:
                    sheet.append(row)

                # Save the changes
                book.save(exportpath)

            # DataFrames and their corresponding sheet names for export data
            export_dataframes_and_sheets = {
                'Relization Item Lines': export_sheet_df1,
                'Relization Mean': export_sheet_df2,
                'Missing Items': export_sheet_df3
            }

            # Loop through the dictionary and append each DataFrame to its corresponding sheet for export data
            for sheet_name, df in export_dataframes_and_sheets.items():
                append_to_excel(df, sheet_name)

            # In[ ]:

            import requests
            import json
            import pandas as pd
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry
            from datetime import datetime

            # Assuming you have pandas imported as pd and fl_df1 is your dataframe

            # Define the base URL and endpoint for sales order details
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Sales%20Order'

            # Define the headers
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807',
                'Content-Type': 'application/json'
            }

            # Function to log output
            def log_output(message):
                self.log_output(message)

            # Setup retry strategy
            retry_strategy = Retry(
                total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504],
                allowed_methods=["GET", "POST", "PUT"]
            )

            # Create a session
            session = requests.Session()
            adapter = HTTPAdapter(max_retries=retry_strategy)
            session.mount("https://", adapter)
            session.mount("http://", adapter)

            # Example: Rename columns in the DataFrame
            sheet_df1.rename(columns={"name": "Sales Order"}, inplace=True)
            sheet_df1.loc[:, 'Date and Time'] = datetime.now()
            sheet_df1.loc[:, 'realization'] = sheet_df1['realization'] / 100

            fl_df1.rename(columns={"name": "name"}, inplace=True)
            fl_df2.rename(columns={
                # Provide the actual column renaming map if required
            }, inplace=True)

            # Iterate through each row in the fl_df1 dataframe
            for index, row in fl_df1.iterrows():
                # Check if name is nan, if so, skip this iteration
                if pd.isnull(row['name']):
                    log_output("Skipping row with NaN name")
                    continue

                # Extract necessary information
                name = row['name']
                realization = row['realization']
                reason_for_hold = row['reason_for_hold']

                # Construct the URL for the specific sales order
                url = f"{base_url}{endpoint}/{name}"

                # Define the payload (body) for the PUT request
                payload = {
                    "realization": realization,
                    "reason_for_hold": reason_for_hold
                }

                # Convert payload to JSON format
                json_payload = json.dumps(payload)

                try:
                    # Send PUT request
                    response = session.put(url, headers=headers, data=json_payload, timeout=10)

                    # Check if request was successful
                    if response.status_code == 200:
                        log_output(f"Successfully updated data for {name}")
                    else:
                        log_output(f"Failed to update data for {name}. Status code: {response.status_code}")
                except requests.exceptions.RequestException as e:
                    log_output(f"Request failed for {name}. Error: {e}")
            self.log_output(f"Realization Completed, Calculating SDR Counts for issue.")
            import requests
            import pandas as pd
            from datetime import datetime, timedelta

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Issue'

            # Calculate the date that is 6 months prior to the current date
            six_months_ago = (datetime.now() - timedelta(days=1 * 30)).strftime('%Y-%m-%d')

            url = base_url + endpoint

            # Define the parameters for the request
            params = {
                'fields': '["name", "issue_generate_date", "custom_sdr_count_last_six_months","issue_details.serial_no"]',
                'limit_start': 0,  # Start from the first record
                'limit_page_length': 1000,  # Request a large number of records per page
                'filters': f'[["custom_sdr_count_last_six_months", "is", "not set"], ["issue_generate_date", ">", "{six_months_ago}"]]'
            }

            # Define the headers if needed
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Update limit_start in params for each iteration
                params['limit_start'] = limit_start

                try:
                    response = requests.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        current_page_data = data['data']
                        all_data.extend(current_page_data)

                        # Check if there are more records
                        if len(current_page_data) < limit_page_length:
                            break  # No more records, exit loop
                        else:
                            limit_start += limit_page_length  # Move to the next page
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Create DataFrame
            issue = pd.json_normalize(all_data)

            import requests
            import pandas as pd
            from datetime import datetime, timedelta

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Service Report'

            # Calculate the date that is 12 months prior to the current date
            twelve_months_ago = (datetime.now() - timedelta(days=365)).strftime('%Y-%m-%d')

            url = base_url + endpoint

            # Define the parameters for the request
            params = {
                'fields': '["mntc_date","serial_number"]',
                'limit_start': 0,  # Start from the first record
                'limit_page_length': 1000,  # Request a large number of records per page
                'filters': f'[["mntc_date", ">", "{twelve_months_ago}"],["job_type", "in", ["SERVICE AND VISIT", "COMPLAINT"]]]'
            }

            # Define the headers if needed
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Update limit_start in params for each iteration
                params['limit_start'] = limit_start

                try:
                    response = requests.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        current_page_data = data['data']
                        all_data.extend(current_page_data)

                        # Check if there are more records
                        if len(current_page_data) < limit_page_length:
                            break  # No more records, exit loop
                        else:
                            limit_start += limit_page_length  # Move to the next page
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Create DataFrame
            s_df = pd.json_normalize(all_data)
            s_df.rename(columns={'serial_number': 'serial_no'}, inplace=True)

            df = pd.merge(issue, s_df, on='serial_no', how='left')

            # Convert columns to datetime
            df['issue_generate_date'] = pd.to_datetime(df['issue_generate_date'])
            df['mntc_date'] = pd.to_datetime(df['mntc_date'])

            # Filter rows
            df_filtered = df[(df['mntc_date'] <= df['issue_generate_date']) &
                             (df['mntc_date'] >= df['issue_generate_date'] - pd.DateOffset(months=6))]

            # Group by 'name' and count 'mntc_date', then rename the column
            count_df = df_filtered.groupby('name').size().reset_index(name='custom_sdr_count_last_six_months')

            # Merge the count back to the filtered DataFrame
            result_df = df_filtered.drop(columns='custom_sdr_count_last_six_months').merge(count_df, on='name')

            final_df = result_df[['name', 'custom_sdr_count_last_six_months']]

            # Drop duplicates based on 'name' and keep only the first occurrence
            final_df = final_df.drop_duplicates(subset='name', keep='first')

            import requests
            import json

            # Assuming you have pandas imported as pd and final_df is your dataframe

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Issue'

            # Define the headers if needed
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807',
                'Content-Type': 'application/json'
            }

            # Iterate through each row in the final_df dataframe
            for index, row in final_df.iterrows():
                # Check if customer_name is nan, if so, skip this iteration
                if pd.isnull(row['name']):
                    self.log_output("Skipping row with NaN customer name")
                    continue

                # Extract necessary information
                name = row['name']
                custom_sdr_count_last_six_months = row['custom_sdr_count_last_six_months']

                # Construct the URL for the specific customer
                url = f"{base_url}{endpoint}/{name}"

                # Define the payload (body) for the PUT request
                payload = {
                    "name": name,
                    "custom_sdr_count_last_six_months": custom_sdr_count_last_six_months
                }

                # Convert payload to JSON format
                json_payload = json.dumps(payload)

                # Send PUT request
                response = requests.put(url, headers=headers, data=json_payload)

                # Check if request was successful
                if response.status_code == 200:
                    self.log_output(f"Successfully updated data for {name}")
                else:
                    self.log_output(f"Failed to update data for {name}. Status code: {response.status_code}")
            self.log_output(f"Issue Completed, Calculating SDR Counts for Warranty.")
            import requests
            import pandas as pd
            from datetime import datetime, timedelta

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Warranty Claim'

            # Calculate the date that is 6 months prior to the current date
            six_months_ago = (datetime.now() - timedelta(days=1 * 30)).strftime('%Y-%m-%d')

            url = base_url + endpoint

            # Define the parameters for the request
            params = {
                'fields': '["name", "complaint_date", "custom_sdr_count_last_six_months","serial_no"]',
                'limit_start': 0,  # Start from the first record
                'limit_page_length': 1000,  # Request a large number of records per page
                'filters': f'[["custom_sdr_count_last_six_months", "is", "not set"], ["complaint_date", ">", "{six_months_ago}"]]'
            }

            # Define the headers if needed
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807'
            }

            # Initialize variables for pagination
            limit_start = 0
            limit_page_length = 1000
            all_data = []

            # Loop to handle pagination
            while True:
                # Update limit_start in params for each iteration
                params['limit_start'] = limit_start

                try:
                    response = requests.get(url, params=params, headers=headers, timeout=10)
                    response.raise_for_status()

                    # Process the response
                    data = response.json()
                    if 'data' in data:
                        current_page_data = data['data']
                        all_data.extend(current_page_data)

                        # Check if there are more records
                        if len(current_page_data) < limit_page_length:
                            break  # No more records, exit loop
                        else:
                            limit_start += limit_page_length  # Move to the next page
                    else:
                        break  # Exit if no data key in response

                except requests.exceptions.RequestException as e:
                    self.log_output(f"Error: {e}")
                    break

            # Create DataFrame
            warranty = pd.json_normalize(all_data)

            df = pd.merge(warranty, s_df, on='serial_no', how='left')

            # Convert columns to datetime
            df['complaint_date'] = pd.to_datetime(df['complaint_date'])
            df['mntc_date'] = pd.to_datetime(df['mntc_date'])

            # Filter rows
            df_filtered = df[(df['mntc_date'] <= df['complaint_date']) &
                             (df['mntc_date'] >= df['complaint_date'] - pd.DateOffset(months=6))]

            # Group by 'name' and count 'mntc_date', then rename the column
            count_df = df_filtered.groupby('name').size().reset_index(name='custom_sdr_count_last_six_months')

            # Merge the count back to the filtered DataFrame
            result_df = df_filtered.drop(columns='custom_sdr_count_last_six_months').merge(count_df, on='name')

            final_df = result_df[['name', 'custom_sdr_count_last_six_months']]

            # Drop duplicates based on 'name' and keep only the first occurrence
            final_df = final_df.drop_duplicates(subset='name', keep='first')

            import requests
            import json

            # Assuming you have pandas imported as pd and final_df is your dataframe

            # Define the base URL and endpoint for stock entry
            base_url = 'https://erpv14.electrolabgroup.com/'
            endpoint = 'api/resource/Warranty Claim'

            # Define the headers if needed
            headers = {
                'Authorization': 'token 3ee8d03949516d0:6baa361266cf807',
                'Content-Type': 'application/json'
            }

            # Iterate through each row in the final_df dataframe
            for index, row in final_df.iterrows():
                # Check if customer_name is nan, if so, skip this iteration
                if pd.isnull(row['name']):
                    self.log_output("Skipping row with NaN customer name")
                    continue

                # Extract necessary information
                name = row['name']
                custom_sdr_count_last_six_months = row['custom_sdr_count_last_six_months']

                # Construct the URL for the specific customer
                url = f"{base_url}{endpoint}/{name}"

                # Define the payload (body) for the PUT request
                payload = {
                    "name": name,
                    "custom_sdr_count_last_six_months": custom_sdr_count_last_six_months
                }

                # Convert payload to JSON format
                json_payload = json.dumps(payload)

                # Send PUT request
                response = requests.put(url, headers=headers, data=json_payload)

                # Check if request was successful
                if response.status_code == 200:
                    self.log_output(f"Successfully updated data for {name}")
                else:
                    self.log_output(f"Failed to update data for {name}. Status code: {response.status_code}")




            
            # Close the session after the loop
            session.close()

            from datetime import datetime

            # Get the current date and time
            current_datetime = datetime.now()

            # Format the datetime object as a string
            formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")

            end_time = time.time()
            excution_time = end_time - start_time
            execution_time = round(excution_time, 2)
            self.log_output(f"Execution Time: {execution_time} seconds")

            # Log output to text area
            self.log_output(f"Completed! Updated at: {formatted_datetime}")

            # Sleep for 2 minutes
            time.sleep(15)

    def clear_text_area(self):
        self.output_text.delete(1.0, tk.END)

    def log_output(self, message):
        self.output_text.insert(tk.END, f"{message}\n")
        self.output_text.see(tk.END)  # Scroll to the bottom of the text area


root = tk.Tk()
app = App(root)
root.mainloop()